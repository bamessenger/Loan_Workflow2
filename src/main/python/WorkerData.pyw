import uuid
import pandas as pd
import numpy as np
import pathlib as p
import win32com.client as win32
from datetime import datetime
from PyQt5.QtWidgets import QMessageBox
from time import sleep
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from PyQt5.QtCore import QObject, pyqtSignal, QThreadPool, QRunnable
from win32com.server.exception import COMException


class WorkerSignals(QObject):
    # Create worker signals
    started = pyqtSignal(str)
    currentStatus = pyqtSignal(str)
    output = pyqtSignal(str)
    tskComplete = pyqtSignal(int)
    completed = pyqtSignal()


class WorkerManager(QObject):
    _workers = {}

    def __init__(self):
        super().__init__()

        # Create a threadpool for workers.
        self.threadpool = QThreadPool()
        self.signals = WorkerSignals()

    def enqueue(self, worker):
        self.threadpool.start(worker)
        self._workers[worker.jobID] = worker

    def notifyCompletion(self, jobID):
        pass


class DataWorker(QRunnable):
    # Worker for the data transfer of Encompass data
    def __init__(self, ePath, wdPath, wrPath):
        super().__init__()
        # create unique identifier for each worker
        self.jobID = str(uuid.uuid4().hex)
        self.efile = ePath
        self.wdfile = wdPath
        self.wrfile = wrPath
        self.signals = WorkerSignals()
        self.msgBox = QMessageBox()

    def run(self):
        self.signals.currentStatus.emit('Running')
        self.fileRead(encompPath=self.efile)
        self.excelWrite(wrkflwDataPath=self.wdfile)
        self.dashData(wrkflwDataPath=self.wdfile, wrkflwRptPath=self.wrfile)

    def fileRead(self, encompPath):
        self.encmpDataAll = pd.read_excel(encompPath, engine='openpyxl')
        self.efileName = p.Path(encompPath).stem
        self.encmpDataAll.columns = self.encmpDataAll.columns.str.replace(
            ' ', '').str.replace('MilestoneDate-', '')
        self.encmpDataAll['LoanStatus'] = np.where(
            self.encmpDataAll.FundingFundsSentDate.isnull(), 'Open', 'Closed')
        self.encmpDataAllAct = self.encmpDataAll.assign(DateType='Actual')
        self.encmpDataAllAct[['ApplicationDate', 'Disclosures',
                              'Processing', 'submittal', 'Approval',
                              'ConditionSubmission', 'ClearToClose',
                              'OS-FinancingContingency',
                              'LockExpirationDate', 'EstClosingDate',
                              'FundingFundsSentDate']] = \
            self.encmpDataAllAct[['ApplicationDate', 'Disclosures',
                                  'Processing', 'submittal', 'Approval',
                                  'ConditionSubmission', 'ClearToClose',
                                  'OS-FinancingContingency',
                                  'LockExpirationDate', 'EstClosingDate',
                                  'FundingFundsSentDate']].apply(pd.to_datetime)
        # Calculate Net working days
        # self.encmpDataAllAct['NetWorkDays'] = np.where(self.encmpDataAllAct['LoanStatus'] == 'Closed',
        #             (np.busday_count(self.encmpDataAllAct['ApplicationDate'].values.astype('datetime64[D]'),
        #                              self.encmpDataAllAct['FundingFundsSentDate'].values.astype('datetime64[D]'))))
        self.signals.output.emit('Create tblEncompassDataAll......Done')
        self.signals.tskComplete.emit(1)
        sleep(2)
        self.encmpDataOpen = self.encmpDataAll[
            self.encmpDataAll.FundingFundsSentDate.isnull()]
        self.encmpDataOpen = self.encmpDataOpen.reset_index()
        self.signals.output.emit('Create tblEncompassOpen......Done')
        self.signals.tskComplete.emit(1)
        sleep(2)
        self.signals.output.emit(self.efileName + ' file read......Done')
        self.signals.tskComplete.emit(1)
        sleep(2)

    def excelWrite(self, wrkflwDataPath):
        self.wdfileName = p.Path(wrkflwDataPath).stem
        self.signals.output.emit('Starting file write to ' +
                                 self.wdfileName + '.....Done')
        self.signals.tskComplete.emit(1)
        wrkbk = load_workbook(wrkflwDataPath)
        # Clean up current sheets in order to create new
        sheetAll = 'tblEncompassAllAct'
        sheetOpen = 'tblEncompassOpen'
        for sheet in wrkbk.sheetnames:
            if sheet == sheetAll:
                wrkbk.remove(wrkbk[sheet])
            elif sheet == sheetOpen:
                wrkbk.remove(wrkbk[sheet])
        self.signals.output.emit('Clean up old data.....Done')
        self.signals.tskComplete.emit(1)
        # Create Excel Writer used to create tables
        writer = pd.ExcelWriter(wrkflwDataPath, mode='a',
                                datetime_format='MM-DD-YYYY', engine='openpyxl')
        writer.book = wrkbk
        # Create tblEncompassAllAct
        self.encmpDataAllAct.to_excel(writer, sheet_name='tblEncompassAllAct',
                                      startcol=1, index=False)
        sheet = wrkbk.get_sheet_by_name('tblEncompassAllAct')
        table = Table(displayName='tblEncompassAllAct',
                      ref='B1:' + get_column_letter(sheet.max_column) + str(
                          sheet.max_row))
        style = TableStyleInfo(name='TableStyleMedium11', showRowStripes=False,
                               showColumnStripes=False)
        table.tableStyleInfo = style
        sheet.add_table(table)
        self.signals.output.emit('Write tblEncompassAllAct to '
                                 + self.wdfileName + '......Done')
        self.signals.tskComplete.emit(1)
        sleep(2)
        # Create tblEncompassOpen
        self.encmpDataOpen.sort_values(by=['ApplicationDate'], inplace=True,
                                       ignore_index=True)
        self.encmpDataOpen['index'] = self.encmpDataOpen.index
        self.encmpDataOpen.to_excel(writer, sheet_name='tblEncompassOpen',
                                    startcol=0, index=False)
        sheet = wrkbk.get_sheet_by_name('tblEncompassOpen')
        table = Table(displayName='tblEncompassOpen',
                      ref='A1:' + get_column_letter(sheet.max_column) + str(
                          sheet.max_row))
        style = TableStyleInfo(name='TableStyleMedium11', showRowStripes=False,
                               showColumnStripes=False)
        table.tableStyleInfo = style
        sheet.add_table(table)
        self.signals.output.emit(
            'Write tblEncompassOpen to ' + self.wdfileName + '......Done')
        self.signals.tskComplete.emit(1)
        wrkbk.save(wrkflwDataPath)
        wrkbk.close()

    def dashData(self, wrkflwDataPath, wrkflwRptPath):
        # Open Workbook up and allow functions to compile
        excel = win32.gencache.EnsureDispatch('Excel.Application')
        workbook = excel.Workbooks.Open(wrkflwDataPath)
        workbook.Save()
        workbook.Close()
        excel.Quit()
        self.signals.output.emit(
            'Open ' + self.wdfileName +
            ' to allow functions to compile......Done')
        self.signals.tskComplete.emit(1)
        wrkbk = load_workbook(wrkflwDataPath)
        # Clean up current sheets in order to create new
        sheetDash = 'tblEncompassAllDash'
        sheetDash2 = 'tblEncompassAllDash2'
        sheetLoanHealth = 'tblEncompassLoanHealth'
        sheetLastComp = 'tblEncompassLastComp'
        sheetTaskList = 'tblEncompassTaskList'
        for sheet in wrkbk.sheetnames:
            if sheet == sheetDash:
                wrkbk.remove(wrkbk[sheet])
            elif sheet == sheetDash2:
                wrkbk.remove(wrkbk[sheet])
            elif sheet == sheetLoanHealth:
                wrkbk.remove(wrkbk[sheet])
            elif sheet == sheetLastComp:
                wrkbk.remove(wrkbk[sheet])
            elif sheet == sheetTaskList:
                wrkbk.remove(wrkbk[sheet])
        # Create Excel Writer used to create tables
        writer = pd.ExcelWriter(wrkflwDataPath, mode='a',
                                datetime_format='MM-DD-YYYY', engine='openpyxl')
        writer.book = wrkbk
        self.encmpDataAllExp = pd.read_excel(wrkflwDataPath, engine='openpyxl',
                                             sheet_name='tblEncompassAllExp')
        self.signals.output.emit('Read tblEncompassAllExp......Done')
        self.signals.tskComplete.emit(1)
        indexNamesExp = self.encmpDataAllExp[(self.encmpDataAllExp[
                                               'LoanNumber'] == 0) |
                                          (self.encmpDataAllExp['LoanNumber']
                                           == " ")].index
        self.encmpDataAllExp.drop(indexNamesExp, inplace=True)
        self.encmpDataAllExp[
             ['ApplicationDate', 'Disclosures', 'Processing', 'submittal',
              'Approval', 'ConditionSubmission', 'ClearToClose',
              'OS-FinancingContingency', 'LockExpirationDate', 'EstClosingDate',
              'FundingFundsSentDate']] = self.encmpDataAllExp[
             ['ApplicationDate', 'Disclosures', 'Processing', 'submittal',
              'Approval', 'ConditionSubmission', 'ClearToClose',
              'OS-FinancingContingency', 'LockExpirationDate', 'EstClosingDate',
              'FundingFundsSentDate']].apply(pd.to_datetime, errors='coerce')
        self.encmpDataAllExp['LoanStatus'] = np.where(
            self.encmpDataAllExp.FundingFundsSentDate.isnull(), 'Open', 'Closed')
        # Create tblEncompassAllDash
        self.encmpDataAllDash = pd.concat(
            [self.encmpDataAllAct, self.encmpDataAllExp])
        self.encmpDataAllDash = self.encmpDataAllDash.melt(
            id_vars=['Company-UsersOrganizationCode', 'LoanOfficer',
                     'LoanProcessor', 'BorrowerLastName', 'LoanNumber',
                     'LoanPurpose', 'LockRequestLoanAmount',
                     'LoanTeamMemberName-UW1-Initial', 'LoanStatus', 'DateType'
                     ], var_name='MilestoneType', value_name='MilestoneDates')
        self.signals.output.emit('Reformat tblEncompassAllDash......Done')
        self.signals.tskComplete.emit(1)
        self.encmpDataAllDash['MilestoneOrder'] = \
            [1 if x == 'Disclosures'
             else 2 if x == 'Processing'
             else 3 if x == 'submittal'
             else 4 if x == 'Approval'
             else 5 if x == 'ConditionSubmission'
             else 6 if x == 'ClearToClose'
             else 99 for x in self.encmpDataAllDash['MilestoneType']]
        self.encmpDataAllDash['MilestoneDates'].replace('', np.nan,
                                                        inplace=True)
        self.encmpDataAllDash.dropna(subset=['LoanNumber'], inplace=True)
        self.encmpDataAllDash['Lookup'] = \
            self.encmpDataAllDash['LoanNumber'].astype('int64').astype(str) + \
            self.encmpDataAllDash['DateType'] + \
            self.encmpDataAllDash['MilestoneType']
        self.signals.output.emit('Create tblEncompassAllDash......Done')
        self.signals.tskComplete.emit(1)
        self.encmpDataAllDash.to_excel(writer, sheet_name='tblEncompassAllDash',
                                       startcol=1, index=False)
        sheet = wrkbk.get_sheet_by_name('tblEncompassAllDash')
        table = Table(displayName='tblEncompassAllDash',
                      ref='B1:' + get_column_letter(sheet.max_column) + str(
                          sheet.max_row))
        style = TableStyleInfo(name='TableStyleMedium11', showRowStripes=False,
                               showColumnStripes=False)
        table.tableStyleInfo = style
        sheet.add_table(table)
        self.signals.output.emit(
            'Write tblEncompassAllDash to ' + self.wdfileName + '......Done')
        self.signals.tskComplete.emit(1)
        self.encmpDataAllDash['Lookup2'] = self.encmpDataAllDash[
                                              'LoanNumber'].astype(
            'int64').astype(str) + self.encmpDataAllDash['MilestoneType']
        self.encmpDataAllDash2 = self.encmpDataAllDash.\
            pivot(index=['Lookup2','LoanOfficer', 'LoanProcessor',
                         'LoanNumber', 'BorrowerLastName', 'MilestoneType',
                         'MilestoneOrder', 'LoanStatus', 'LoanPurpose'],
                  columns='DateType',
                  values='MilestoneDates').\
                           rename_axis(None, axis=1).reset_index()
        self.encmpDataAllDash2.fillna('', inplace=True)
        self.encmpDataAllDash2[['Expected', 'Actual']] = \
            self.encmpDataAllDash2[['Expected', 'Actual']].\
                apply(pd.to_datetime, errors='coerce')
        self.encmpDataAllDash2.sort_values(by=['Expected', 'MilestoneOrder'],
                                           ascending=[True, True])
        self.signals.output.emit('Create tblEncompassAllDash2......Done')
        self.signals.tskComplete.emit(1)
        self.encmpDataAllDash2.to_excel(writer,
                                        sheet_name='tblEncompassAllDash2',
                                        startcol=1, index=False)
        sheet = wrkbk.get_sheet_by_name('tblEncompassAllDash2')
        table = Table(displayName='tblEncompassAllDash2',
                      ref='B1:' + get_column_letter(sheet.max_column) + str(
                          sheet.max_row))
        style = TableStyleInfo(name='TableStyleMedium11', showRowStripes=False,
                               showColumnStripes=False)
        table.tableStyleInfo = style
        sheet.add_table(table)
        self.signals.output.emit(
            'Write tblEncompassAllDash2 to ' + self.wdfileName + '......Done')
        self.signals.tskComplete.emit(1)
        # Create tblEncompassLoanHealth
        self.encmpDataLoanHlth = pd.concat(
            [self.encmpDataAllAct, self.encmpDataAllExp])
        self.encmpDataLoanHlth.to_excel(writer, sheet_name='tblEncompassLoanHealth',
                                       startcol=1, index=False)
        sheet = wrkbk.get_sheet_by_name('tblEncompassLoanHealth')
        table = Table(displayName='tblEncompassLoanHealth',
                      ref='B1:' + get_column_letter(sheet.max_column) + str(
                          sheet.max_row))
        style = TableStyleInfo(name='TableStyleMedium11', showRowStripes=False,
                               showColumnStripes=False)
        table.tableStyleInfo = style
        sheet.add_table(table)
        self.signals.output.emit('Create tblEncompassLoanHealth......Done')
        self.signals.tskComplete.emit(1)
        wrkbk.save(wrkflwDataPath)
        wrkbk.close()
        # Create tblEncompassLastComp
        self.encmpDataLastComp = self.encmpDataAllDash[(self.encmpDataAllDash['MilestoneOrder'] != 99) &
                                                       (self.encmpDataAllDash['LoanStatus'] == 'Open') &
                                                       (self.encmpDataAllDash['DateType'] == 'Actual')]
        self.encmpDataLastComp.dropna(subset=['MilestoneDates'], inplace=True)
        self.encmpDataLastComp = self.encmpDataLastComp.sort_values(['LoanNumber', 'MilestoneOrder'],
                                 ascending=[True, False]).drop_duplicates(['LoanNumber']).reset_index(drop=True)
        self.encmpDataLastComp = self.encmpDataLastComp.drop(columns=['Company-UsersOrganizationCode', 'LoanOfficer',
                                                                      'LoanProcessor','BorrowerLastName', 'LoanPurpose',
                                                                      'LockRequestLoanAmount', 'LoanStatus', 'DateType',
                                                                      'MilestoneOrder','LoanTeamMemberName-UW1-Initial',
                                                                      'Lookup', 'Lookup2'])
        self.encmpDataLastComp.rename(columns= {'MilestoneType':'LastCompletedMilestone',
                                                'MilestoneDates':'LastCompletedMilestoneDate'}, inplace=True)
        self.encmpDataLastComp.to_excel(writer, sheet_name='tblEncompassLastComp',
                                        startcol=1, index=False)
        sheet = wrkbk.get_sheet_by_name('tblEncompassLastComp')
        table = Table(displayName='tblEncompassLastComp',
                      ref='B1:' + get_column_letter(sheet.max_column) + str(
                          sheet.max_row))
        style = TableStyleInfo(name='TableStyleMedium11', showRowStripes=False,
                               showColumnStripes=False)
        table.tableStyleInfo = style
        sheet.add_table(table)
        self.signals.output.emit('Create tblEncompassLastComp......Done')
        self.signals.tskComplete.emit(1)
        # Create tblEncompassTaskList
        self.encmpDataTaskList = self.encmpDataAllDash2[(self.encmpDataAllDash2['MilestoneOrder'] != 99) &
                                                        (self.encmpDataAllDash2['LoanStatus'] == 'Open')]
        self.encmpDataTaskList = pd.merge(left=self.encmpDataTaskList, how='left',
                                          right=self.encmpDataLastComp, on=['LoanNumber'])
        self.encmpDataTaskList.to_excel(writer, sheet_name='tblEncompassTaskList',
                                        startcol=1, index=False)
        sheet = wrkbk.get_sheet_by_name('tblEncompassTaskList')
        table = Table(displayName='tblEncompassTaskList',
                      ref='B1:' + get_column_letter(sheet.max_column) + str(
                          sheet.max_row))
        style = TableStyleInfo(name='TableStyleMedium11', showRowStripes=False,
                               showColumnStripes=False)
        table.tableStyleInfo = style
        sheet.add_table(table)
        self.signals.output.emit('Create tblEncompassTaskList......Done')
        self.signals.tskComplete.emit(1)
        wrkbk.save(wrkflwDataPath)
        wrkbk.close()
        # Open DailyWorkflowRpting.xlsx to be fully updated with DailyWorkflowData.xlsx data
        # program will pause until Queries are done refreshing
        excel = win32.gencache.EnsureDispatch('Excel.Application')
        workbookRpt = excel.Workbooks.Open(wrkflwRptPath)
        workbookRpt.RefreshAll()
        excel.CalculateUntilAsyncQueriesDone()
        self.signals.output.emit('Refresh Queries......Done')
        self.signals.tskComplete.emit(1)
        workbookRpt.Save()
        workbookRpt.Close()
        excel.Quit()
        # Datestamp DailyWorkflowRpting workbook
        excel = win32.gencache.EnsureDispatch('Excel.Application')
        workbookRpt = excel.Workbooks.Open(wrkflwRptPath)
        self.wrfileName = p.Path(wrkflwRptPath).stem
        wrkShtSettings = workbookRpt.Worksheets('settings')
        wrkShtSettings.Cells(2, 5).Value = datetime.today()
        self.signals.output.emit('Timestamp ' + self.wdfileName + '......Done')
        self.signals.tskComplete.emit(1)
        workbookRpt.Save()
        workbookRpt.Close()
        excel.Quit()
        self.signals.output.emit(
            self.wdfileName + ' saved and closed......Done')
        self.signals.tskComplete.emit(1)
        sleep(1)
        self.signals.currentStatus.emit('Idle')
        self.signals.completed.emit()

